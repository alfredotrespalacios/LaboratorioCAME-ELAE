from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from came.data.monthly_store import (
    allocate_ready_package_directory,
    create_monthly_package,
    create_stored_monthly_package,
    get_package_spec,
    store_monthly_package,
)
from came.exports import build_excel, build_pdf
from came.report import build_executive_prompt, make_package


def test_excel_contains_standard_sheets_and_timezone_safe_dates() -> None:
    data = pd.DataFrame(
        {"datetime": pd.date_range("2024-01-01", periods=2, tz="UTC"), "value": [1, 2]}
    )
    content = build_excel(data=data, summary={"Promedio": 1.5}, methodology=["Prueba"])
    workbook = load_workbook(BytesIO(content), read_only=True)
    assert workbook.sheetnames[:5] == ["Datos", "Resumen", "Parámetros", "Metodología", "Cobertura"]


def test_pdf_is_generated() -> None:
    content = build_pdf(
        title="Prueba",
        subtitle="Cobertura",
        indicators={"Valor": 1},
        tables={"Datos": pd.DataFrame({"a": [1]})},
    )
    assert content.startswith(b"%PDF")
    assert len(content) > 1000


def test_executive_prompt_caps_news_and_questions_and_contains_packages() -> None:
    package = make_package(
        module="1",
        title="Precio",
        period="2024",
        source="XM",
        unit="COP/kWh",
        configuration={},
        indicators={"promedio": 100},
        methodology=["Promedio"],
    )
    prompt = build_executive_prompt(
        [package],
        audience="Directivos",
        tone="Ejecutivo",
        length="2 páginas",
        technical_level="Intermedio",
        news=[{"titulo": str(index)} for index in range(5)],
        questions=[f"Q{index}" for index in range(6)],
    )
    assert '"title": "Precio"' in prompt
    assert "Q3" in prompt
    assert "Q4" not in prompt
    assert '"titulo": "2"' in prompt
    assert '"titulo": "3"' not in prompt


def test_app_declares_all_19_numbered_pages() -> None:
    text = (Path(__file__).resolve().parents[1] / "app.py").read_text(encoding="utf-8")
    for number in range(1, 20):
        assert f'title="{number}.' in text


def test_each_visible_module_has_its_own_page_file() -> None:
    pages = Path(__file__).resolve().parents[1] / "src" / "came" / "ui" / "pages"
    expected = {
        "price_spot.py",
        "demand_national.py",
        "generation_technology.py",
        "generation_resource.py",
        "xm_explorer.py",
        "base_integrated.py",
        "energy_balance.py",
        "offer_curve.py",
        "spain.py",
        "chile.py",
        "modeling_forecast.py",
        "sarima_garch.py",
        "portfolio_montecarlo.py",
        "introduction.py",
        "case_studies.py",
        "executive_report.py",
        "data_maintenance.py",
    }
    assert expected.issubset({path.name for path in pages.glob("*.py")})


def test_introduction_precedes_colombia_and_case_studies_are_named() -> None:
    text = (Path(__file__).resolve().parents[1] / "app.py").read_text(encoding="utf-8")
    assert text.index('"Inicio"') < text.index('"Colombia"')
    assert 'title="Introducción"' in text
    assert 'default=True' in text
    assert '"Casos de estudio"' in text


def test_integrated_module_points_to_data_maintenance() -> None:
    page = (
        Path(__file__).resolve().parents[1] / "src" / "came" / "ui" / "pages" / "base_integrated.py"
    ).read_text(encoding="utf-8")
    assert "load_default_monthly" in page
    assert "Mantenimiento → Mantenimiento de datos" in page
    assert "Construir consulta temporal" in page
    assert "Streamlit no modifica GitHub automáticamente" in page
    assert "administrador" not in page.lower()


def test_data_maintenance_explains_the_complete_update_flow() -> None:
    page = (
        Path(__file__).resolve().parents[1]
        / "src"
        / "came"
        / "ui"
        / "pages"
        / "data_maintenance.py"
    ).read_text(encoding="utf-8")
    spec = get_package_spec("COL")
    assert spec.relative_directory.as_posix() == "datos_por_defecto/colombia"
    assert spec.parquet_name == "Base_integrada_mensual.parquet"
    assert spec.catalog_name == "Catalogo_Base_integrada.xlsx"
    assert spec.metadata_name == "Fecha_actualizacion_Base_integrada.json"
    assert "Agregar meses faltantes" in page
    assert "Construir la primera base" in page
    assert "Recalcular un periodo" in page
    assert "Paquete listo" in page
    assert "datos_por_defecto/colombia/" in page
    assert "tres intentos automáticos" in page
    assert "la aplicación nunca lo reemplaza automáticamente" in page
    assert "administrador" not in page.lower()


def test_data_maintenance_shows_downloads_from_the_stored_package(tmp_path) -> None:
    data = pd.DataFrame(
        [
            {
                "datetime": pd.Timestamp("2024-01-01", tz="UTC"),
                "country": "COL",
                "family": "Mercado",
                "level": "Sistema",
                "entity_code": "SIN",
                "entity_name": "Colombia",
                "variable": "Precio",
                "unit": "COP/kWh",
                "value": 100.0,
                "source": "XM",
                "dataset": "Prueba/Sistema",
                "aggregation": "Promedio mensual",
                "series_id": "col_precio_prueba",
                "series_name": "Precio de prueba",
                "catalog_date": "2026-08-08",
            }
        ]
    )
    package = create_monthly_package(data, "COL", reference="2024-02-15")
    stored = store_monthly_package(package, tmp_path / "package")
    script = f'''
import pandas as pd
from came.ui.pages.data_maintenance import _show_result

_show_result(
    {{
        "status": pd.DataFrame(),
        "warnings": [],
        "errors": [],
        "package_directory": {str(stored.directory)!r},
        "country": "COL",
    }},
    "stored_package_test",
)
'''

    app = AppTest.from_string(script, default_timeout=30).run()

    assert not app.exception
    labels = [button.label for button in app.get("download_button")]
    assert "Descargar ZIP listo para GitHub" in labels
    assert package.spec.parquet_name not in labels

    app.run()
    labels_after_rerun = [button.label for button in app.get("download_button")]
    assert "Descargar ZIP listo para GitHub" in labels_after_rerun

    individual = next(
        checkbox
        for checkbox in app.checkbox
        if checkbox.label == "Necesito descargar también un archivo individual"
    )
    individual.check().run()
    labels_with_individual = [button.label for button in app.get("download_button")]
    assert f"Descargar {package.spec.parquet_name}" in labels_with_individual


def test_data_maintenance_recovers_latest_zip_in_a_fresh_session(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setenv("CAME_RUNTIME_STORAGE", str(tmp_path / "runtime"))
    data = pd.DataFrame(
        [
            {
                "datetime": pd.Timestamp("2024-01-01", tz="UTC"),
                "country": "COL",
                "family": "Mercado",
                "level": "Sistema",
                "entity_code": "SIN",
                "entity_name": "Colombia",
                "variable": "Precio",
                "unit": "COP/kWh",
                "value": 100.0,
                "source": "XM",
                "dataset": "Prueba/Sistema",
                "aggregation": "Promedio mensual",
                "series_id": "col_precio_prueba",
                "series_name": "Precio de prueba",
                "catalog_date": "2026-08-08",
            }
        ]
    )
    output = allocate_ready_package_directory("COL", "fresh-session")
    create_stored_monthly_package(data, "COL", output, reference="2024-02-15")
    script = """
from came.ui.pages.data_maintenance import _show_result, _state_or_latest

state = _state_or_latest("brand_new_session", "COL")
assert state is not None
assert state["recovered"] is True
_show_result(state, "fresh_session_package")
"""

    app = AppTest.from_string(script, default_timeout=30).run()

    assert not app.exception
    assert "Descargar ZIP listo para GitHub" in [
        button.label for button in app.get("download_button")
    ]


def test_first_base_flow_finishes_with_zip_and_a_new_session_recovers_it(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setenv("CAME_RUNTIME_STORAGE", str(tmp_path / "runtime"))
    script = """
from types import SimpleNamespace

import pandas as pd
import streamlit as st
import came.ui.pages.data_maintenance as page
from came.data.maintenance import BuildResult
from came.data.monthly_store import LONG_COLUMNS

row = {
    "datetime": pd.Timestamp("2024-01-01", tz="UTC"),
    "country": "COL",
    "family": "Mercado",
    "level": "Sistema",
    "entity_code": "SIN",
    "entity_name": "Colombia",
    "variable": "Precio",
    "unit": "COP/kWh",
    "value": 100.0,
    "source": "XM",
    "dataset": "Prueba/Sistema",
    "aggregation": "Promedio mensual",
    "series_id": "col_precio_prueba",
    "series_name": "Precio de prueba",
    "catalog_date": "2026-08-08",
}

class FakeBuilder:
    def __init__(self, *args, **kwargs):
        self.checkpoints = SimpleNamespace(directory=None)

    def clear_checkpoints(self):
        pass

    def build(self, *args, **kwargs):
        st.session_state["captured_selected_options"] = kwargs.get("selected_options")
        return BuildResult(
            country="COL",
            data=pd.DataFrame([row]),
            status=pd.DataFrame({"Fuente": ["XM"], "Estado": ["Aprobado"]}),
        )

page.ColombiaMonthlyBuilder = FakeBuilder
page._load_existing = lambda country: (pd.DataFrame(columns=LONG_COLUMNS), {}, None)
page.page_data_maintenance(1)
"""
    app = AppTest.from_string(script, default_timeout=30).run()
    operation = next(radio for radio in app.radio if radio.label == "Operación para Colombia")
    app = operation.set_value("Construir la primera base").run()
    selection = next(
        multiselect
        for multiselect in app.multiselect
        if multiselect.label == "Canasta CAME preseleccionada"
    )
    app = selection.set_value(["demand"]).run()
    confirmation = next(
        checkbox
        for checkbox in app.checkbox
        if checkbox.label
        == "Entiendo que la operación puede tardar y mantendré abierta esta pestaña."
    )
    confirmation.check().run()
    run_button = next(
        button for button in app.button if button.label == "Reanudar o iniciar construcción"
    )
    run_button.click().run()

    assert not app.exception
    assert app.session_state["captured_selected_options"] == {"demand"}
    assert "Descargar ZIP listo para GitHub" in [
        button.label for button in app.get("download_button")
    ]
    rendered_text = "\n".join(
        str(element.value)
        for element_type in ("markdown", "info", "success", "caption")
        for element in app.get(element_type)
    )
    for phase in ("0/5", "1/5", "2/5", "3/5", "4/5", "5/5"):
        assert phase in rendered_text

    fresh_app = AppTest.from_string(script, default_timeout=30).run()
    assert not fresh_app.exception
    assert "Descargar ZIP listo para GitHub" in [
        button.label for button in fresh_app.get("download_button")
    ]


def test_first_base_initial_date_allows_the_end_of_the_previous_year(monkeypatch) -> None:
    import came.ui.pages.data_maintenance as page

    calls: list[tuple[str, dict[str, object]]] = []

    def fake_date_input(label: str, **kwargs):
        calls.append((label, kwargs))
        return kwargs["value"]

    monkeypatch.setattr(page.st, "date_input", fake_date_input)
    page._period_controls(
        "COL",
        page.OPERATION_BUILD,
        pd.DataFrame(),
        default_start=date(2000, 1, 1),
        key="date_range_test",
    )

    first_call = next(call for call in calls if call[0] == "Fecha inicial de la historia")
    assert first_call[1]["min_value"] == date(2000, 1, 1)
    assert first_call[1]["max_value"] == date(date.today().year - 1, 12, 31)


def test_data_maintenance_explains_why_packaging_did_not_start() -> None:
    script = """
import pandas as pd
from came.data.maintenance import BuildResult
from came.ui.pages.data_maintenance import _create_download_package, _show_result

result = BuildResult(
    country="COL",
    data=pd.DataFrame({"value": [1.0]}),
    status=pd.DataFrame(),
    errors=["XM respondió 502"],
)
package = _create_download_package(
    result,
    country="COL",
    build_id="blocking-error",
    additional_sheets={},
    build_notes=[],
)
_show_result(
    {
        "status": result.status,
        "warnings": result.warnings,
        "errors": result.errors,
        "package_directory": None,
        "country": "COL",
    },
    "blocking_error",
)
"""

    app = AppTest.from_string(script, default_timeout=30).run()

    assert not app.exception
    rendered_text = "\n".join(
        str(element.value)
        for element_type in ("markdown", "caption", "error")
        for element in app.get(element_type)
    )
    assert "0/5" in rendered_text
    for phase in ("1/5", "2/5", "3/5", "4/5", "5/5"):
        assert phase in rendered_text
    assert "El empaquetado NO comenzó" in rendered_text
    assert not app.get("download_button")


def test_result_is_saved_for_the_report_only_after_clicking_the_button() -> None:
    script = """
import pandas as pd
from came.ui.components import export_and_collect

export_and_collect(
    module="Prueba",
    title="Resultado de prueba",
    data=pd.DataFrame({"value": [1.0]}),
    indicators={"promedio": 1.0},
    parameters={},
    methodology=["Prueba"],
    source="Prueba",
    unit="unidad",
    period="2026",
    key="manual_report_test",
)
"""
    app = AppTest.from_string(script, default_timeout=30).run()
    assert not app.exception
    assert "report_packages" not in app.session_state.filtered_state
    save = next(button for button in app.button if "Guardar resultado" in button.label)
    save.click().run()
    assert len(app.session_state["report_packages"]) == 1


def test_generation_resource_excel_includes_all_monthly_history() -> None:
    page = (
        Path(__file__).resolve().parents[1]
        / "src"
        / "came"
        / "ui"
        / "pages"
        / "generation_resource.py"
    ).read_text(encoding="utf-8")
    assert '"Histórico recursos"' in page
    assert '"Histórico empresas"' in page
    assert '"Histórico tecnología"' in page
    assert '"Catálogo recursos"' in page
    assert '"Validación totales"' in page
    assert "todos los recursos consultados" in page


def test_streamlit_default_page_smoke(monkeypatch) -> None:
    monkeypatch.setenv("CAME_DEV_MODE", "1")
    app_path = Path(__file__).resolve().parents[1] / "app.py"
    app = AppTest.from_file(str(app_path), default_timeout=30).run()
    assert not app.exception
    assert app.title[0].value == "Laboratorio CAME"
    assert app.header[0].value == "Introducción"
    assert "Esta aplicación es propiedad de ELAE" in app.info[0].value


@pytest.mark.parametrize(
    "module,function,arguments",
    [
        ("base_integrated", "page_integrated", "45"),
        ("spain", "page_spain", "45"),
        ("chile", "page_chile", "45"),
        ("modeling_forecast", "page_modeling", ""),
        ("sarima_garch", "page_volatility", ""),
        ("portfolio_montecarlo", "page_portfolio", ""),
        ("data_maintenance", "page_data_maintenance", "45"),
    ],
)
def test_new_monthly_pages_render_without_live_queries(
    module: str, function: str, arguments: str
) -> None:
    script = f"from came.ui.pages.{module} import {function}\n{function}({arguments})\n"
    app = AppTest.from_string(script, default_timeout=30).run()
    assert not app.exception
